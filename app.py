from fastapi import FastAPI
from fastapi.responses import FileResponse
from pydantic import BaseModel
from openpyxl import load_workbook
from datetime import date
import shutil
import os
import re

app = FastAPI()

@app.get("/")
def root():
    return {"status": "Rental API is running"}

class RentalPayload(BaseModel):
    unit_number: str
    complex_name: str
    property_address: str
    occupation_date: str
    marketing_amount: float
    owner_pocket_amount: float
    lease_term: int
    placement_fee: float
    management_fee: float
    tenrisk: bool
    renewal_fee: float
    placement_term: int

@app.post("/export-rental-calc")
def export_rental_calc(data: RentalPayload):
    base_dir = os.path.dirname(os.path.abspath(__file__))
    template = os.path.join(base_dir, "Rental Calc Template.xlsx")

    output_dir = os.path.join(base_dir, "generated_files")
    os.makedirs(output_dir, exist_ok=True)

    safe_address = re.sub(r"[^A-Za-z0-9_-]+", "_", data.property_address).strip("_")
    output_path = os.path.join(output_dir, f"Rental_Calc_{safe_address}.xlsx")

    shutil.copy(template, output_path)

    wb = load_workbook(output_path)
    ws = wb.active

    ws["B4"] = data.unit_number
    ws["B5"] = data.complex_name
    ws["B6"] = data.property_address
    ws["B7"] = date.today().strftime("%d %B %Y")
    ws["B8"] = data.occupation_date

    ws["D11"] = data.marketing_amount
    ws["D12"] = data.owner_pocket_amount
    ws["F11"] = data.lease_term

    ws["B14"] = data.placement_fee
    ws["B15"] = data.management_fee
    ws["B16"] = 0.031 if data.tenrisk else 0
    ws["B20"] = data.renewal_fee

    ws["D22"] = data.placement_term

    wb.save(output_path)

    return FileResponse(
        output_path,
        filename=os.path.basename(output_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )